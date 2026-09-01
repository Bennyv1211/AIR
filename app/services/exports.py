from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import BusinessAssumptions, Recommendation

HEADER_FILL = PatternFill(fill_type="solid", fgColor="E8F1EC")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="F7F1E4")
PRIORITY_FILLS = {
    "critical": PatternFill(fill_type="solid", fgColor="F7D7D7"),
    "high": PatternFill(fill_type="solid", fgColor="FDE5C3"),
    "medium": PatternFill(fill_type="solid", fgColor="DCEAFF"),
    "low": PatternFill(fill_type="solid", fgColor="DDEEDB"),
}
PO_IMPORT_HEADERS = ["PUNO", "ITNO", "ORQA", "PUPR"]
PO_IMPORT_DESCRIPTIONS = [
    "*Purchase order number(10)",
    "*Item number(15)",
    "*Ordered quantity - alternate U/M(17)",
    "Purchase price(19)",
]


def build_recommendations_workbook(
    recommendations: list[Recommendation],
    assumptions: BusinessAssumptions,
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    detail_sheet = workbook.create_sheet("Verdict Detail")
    assumptions_sheet = workbook.create_sheet("Assumptions & Inputs")
    po_import_sheet = workbook.create_sheet("PO Import Template")

    _build_summary_sheet(summary_sheet, recommendations)
    _build_detail_sheet(detail_sheet, recommendations)
    _build_assumptions_sheet(assumptions_sheet, assumptions)
    _build_po_import_sheet(po_import_sheet, recommendations)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_summary_sheet(sheet, recommendations: list[Recommendation]) -> None:
    sheet["A1"] = "AIR Verdict Summary"
    sheet["A1"].font = Font(bold=True, size=16)

    total_items = len(recommendations)
    reorder_items = sum(1 for item in recommendations if item.needs_reorder)
    counts = {
        "critical": sum(1 for item in recommendations if item.priority == "critical"),
        "high": sum(1 for item in recommendations if item.priority == "high"),
        "medium": sum(1 for item in recommendations if item.priority == "medium"),
        "low": sum(1 for item in recommendations if item.priority == "low"),
    }

    summary_rows = [
        ("Total items", total_items),
        ("Need reorder", reorder_items),
        ("Critical priority", counts["critical"]),
        ("High priority", counts["high"]),
        ("Medium priority", counts["medium"]),
        ("Low priority", counts["low"]),
    ]

    for index, (label, value) in enumerate(summary_rows, start=3):
        sheet[f"A{index}"] = label
        sheet[f"B{index}"] = value
        sheet[f"A{index}"].font = Font(bold=True)
        sheet[f"A{index}"].fill = SUMMARY_FILL
        sheet[f"B{index}"].fill = SUMMARY_FILL

    sheet["D3"] = "Top urgent items"
    sheet["D3"].font = Font(bold=True)
    urgent_headers = ["SKU Number", "Name", "Priority", "Suggested Order", "Stockout Date"]
    for column_index, header in enumerate(urgent_headers, start=4):
        cell = sheet.cell(row=4, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    urgent_items = [item for item in recommendations if item.needs_reorder][:10]
    for row_index, item in enumerate(urgent_items, start=5):
        sheet[f"D{row_index}"] = item.sku
        sheet[f"E{row_index}"] = item.name
        sheet[f"F{row_index}"] = item.priority
        sheet[f"G{row_index}"] = item.recommended_order_qty
        sheet[f"H{row_index}"] = item.projected_stockout_date.isoformat() if item.projected_stockout_date else ""
        sheet[f"F{row_index}"].fill = PRIORITY_FILLS[item.priority]

    for column, width in {
        "A": 22,
        "B": 14,
        "D": 16,
        "E": 28,
        "F": 14,
        "G": 18,
        "H": 16,
    }.items():
        sheet.column_dimensions[column].width = width


def _build_detail_sheet(sheet, recommendations: list[Recommendation]) -> None:
    headers = [
        "SKU Number",
        "Name",
        "Current Stock",
        "Reorder Point",
        "Target Stock",
        "Suggested Order",
        "Planned Order Date",
        "Planned Delivery Date",
        "Priority",
        "Stockout Date",
        "Demand Source",
        "Explanation",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for item in recommendations:
        sheet.append(
            [
                item.sku,
                item.name,
                item.current_stock,
                item.reorder_point,
                item.target_stock,
                item.recommended_order_qty,
                item.planned_order_date.isoformat() if item.planned_order_date else "",
                item.planned_delivery_date.isoformat() if item.planned_delivery_date else "",
                item.priority,
                item.projected_stockout_date.isoformat() if item.projected_stockout_date else "",
                item.demand_source,
                item.explanation,
            ]
        )
        priority_cell = sheet.cell(row=sheet.max_row, column=9)
        priority_cell.fill = PRIORITY_FILLS[item.priority]

    widths = {
        "A": 16,
        "B": 30,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 18,
        "I": 12,
        "J": 16,
        "K": 18,
        "L": 80,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, max_col=12):
        row[11].alignment = Alignment(wrap_text=True, vertical="top")


def _build_assumptions_sheet(sheet, assumptions: BusinessAssumptions) -> None:
    sheet["A1"] = "AIR Planning Assumptions"
    sheet["A1"].font = Font(bold=True, size=16)

    rows = [
        ("Shipping days per week", assumptions.shipping_days_per_week or ""),
        ("Order days", ", ".join(assumptions.order_days) if assumptions.order_days else ""),
        ("Arrival days", ", ".join(assumptions.arrival_days) if assumptions.arrival_days else ""),
        ("Default spoilage days", assumptions.default_spoilage_days if assumptions.default_spoilage_days is not None else ""),
        ("Produce spoilage days", assumptions.produce_spoilage_days if assumptions.produce_spoilage_days is not None else ""),
        ("Herb spoilage days", assumptions.herb_spoilage_days if assumptions.herb_spoilage_days is not None else ""),
        ("Additional notes", assumptions.additional_notes or ""),
    ]

    sheet["A3"] = "Input"
    sheet["B3"] = "Value"
    for cell in sheet[3]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row_index, (label, value) in enumerate(rows, start=4):
        sheet[f"A{row_index}"] = label
        sheet[f"B{row_index}"] = value
        sheet[f"A{row_index}"].font = Font(bold=True)
        sheet[f"A{row_index}"].fill = SUMMARY_FILL

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 80

    for row in sheet.iter_rows(min_row=4, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def _build_po_import_sheet(sheet, recommendations: list[Recommendation]) -> None:
    """Create an M3-style PO import tab from the supplied upload template."""
    sheet.append(PO_IMPORT_HEADERS)
    sheet.append(PO_IMPORT_DESCRIPTIONS)

    for item in recommendations:
        if item.needs_reorder and item.recommended_order_qty > 0:
            # Leave PO number and purchase price blank so this tab can be completed upstream.
            sheet.append(["", item.sku, item.recommended_order_qty, ""])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for cell in sheet[2]:
        cell.font = Font(italic=True, color="5A6A63")

    for column, width in {"A": 30, "B": 20, "C": 34, "D": 22}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A3"
