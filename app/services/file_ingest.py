from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from app.models import (
    DatasetImportMetadata,
    FileMappingPreview,
    FilePreviewResponse,
    ReplenishmentRecord,
    UsageRecord,
)
from app.services.openai_semantic import infer_semantic_headers, openai_is_configured

DEFAULT_LEAD_TIME_DAYS = 7
DEFAULT_SAFETY_STOCK = 0
DEFAULT_MIN_ORDER_QTY = 0

HEADER_ALIASES = {
    "sku": {
        "sku",
        "item",
        "itemcode",
        "itemnumber",
        "productcode",
        "stockcode",
        "partnumber",
        "partno",
    },
    "name": {
        "name",
        "description",
        "itemname",
        "productname",
        "itemdescription",
        "productdescription",
    },
    "current_stock": {
        "currentstock",
        "stock",
        "onhand",
        "qoh",
        "qtyonhand",
        "available",
        "inventory",
        "instock",
    },
    "daily_demand": {
        "dailydemand",
        "avgdailydemand",
        "dailyusage",
        "dailyvelocity",
        "avgsalesday",
        "avgdailyunits",
        "dailyrunrate",
        "dailyaveragesales",
        "averagedailysales",
    },
    "lead_time_days": {
        "leadtime",
        "leadtimedays",
        "leadtimeindays",
        "supplierleadtime",
        "vendorleadtime",
        "ltdays",
    },
    "safety_stock": {
        "safetystock",
        "bufferstock",
        "buffer",
        "minimumstock",
        "minsafetystock",
        "stockbuffer",
    },
    "min_order_qty": {
        "minorderqty",
        "minimumorderqty",
        "minimumorderquantity",
        "moq",
        "casepack",
        "packsize",
        "ordermultiple",
    },
    "incoming_stock": {
        "onorder",
        "incoming",
        "arriving",
        "intransit",
        "openpo",
        "purchaseorder",
        "outstandingpo",
        "qtyonorder",
    },
    "supplier_code": {
        "supplier",
        "suppliercode",
        "suppliernumber",
        "vendor",
        "vendorcode",
        "vendornumber",
        "vendorid",
    },
}

USAGE_HEADER_ALIASES = {
    "sku": HEADER_ALIASES["sku"],
    "usage_date": {
        "date",
        "usagedate",
        "transactiondate",
        "day",
        "shipdate",
        "salesdate",
        "activitydate",
    },
    "units_used": {
        "usage",
        "dailyusage",
        "unitsused",
        "unitssold",
        "sold",
        "sales",
        "qtysold",
        "quantitysold",
        "consumption",
        "dailyaveragesales",
        "averagedailysales",
    },
}

WEEKLY_DEMAND_PATTERNS = (
    (re.compile(r"^(last|past)?1wk$"), 7),
    (re.compile(r"^(last|past)?1week$"), 7),
    (re.compile(r"^week1$"), 7),
    (re.compile(r"^wk1$"), 7),
    (re.compile(r"^1weeksales$"), 7),
    (re.compile(r"^1wksales$"), 7),
    (re.compile(r"^lastweek$"), 7),
    (re.compile(r"^lastweeksales$"), 7),
    (re.compile(r"^saleslastweek$"), 7),
    (re.compile(r"^2wk$"), 14),
    (re.compile(r"^2week$"), 14),
    (re.compile(r"^2weeksales$"), 14),
    (re.compile(r"^2wksales$"), 14),
    (re.compile(r"^3wk$"), 21),
    (re.compile(r"^3week$"), 21),
    (re.compile(r"^3weeksales$"), 21),
    (re.compile(r"^3wksales$"), 21),
    (re.compile(r"^4wk$"), 28),
    (re.compile(r"^4week$"), 28),
    (re.compile(r"^4weeksales$"), 28),
    (re.compile(r"^4wksales$"), 28),
)

WEEKDAY_INDEXES = {
    "mon": 0,
    "monday": 0,
    "tue": 1,
    "tues": 1,
    "tuesday": 1,
    "wed": 2,
    "wednesday": 2,
    "thu": 3,
    "thur": 3,
    "thurs": 3,
    "thursday": 3,
    "fri": 4,
    "friday": 4,
    "sat": 5,
    "saturday": 5,
    "sun": 6,
    "sunday": 6,
}

WEEKLY_USAGE_PATTERN = re.compile(
    r"^(?P<week>lw|lastweek|cw|currentweek)(?P<day>mon|monday|tue|tues|tuesday|wed|wednesday|thu|thur|thurs|thursday|fri|friday|sat|saturday|sun|sunday)$"
)


@dataclass
class HeaderMatch:
    canonical_field: str
    source_header: str
    score: float


def parse_replenishment_file(filename: str, content: bytes) -> list[ReplenishmentRecord]:
    records, _ = parse_replenishment_file_with_metadata(filename, content)
    return records


def parse_replenishment_file_with_metadata(
    filename: str,
    content: bytes,
) -> tuple[list[ReplenishmentRecord], DatasetImportMetadata]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_rows(_csv_rows(content), "CSV")
    if suffix == ".xlsx":
        return _parse_rows(_xlsx_rows(content), "Excel")
    raise ValueError("Please upload a .csv or .xlsx file.")


def parse_usage_file(filename: str, content: bytes) -> list[UsageRecord]:
    records, _ = parse_usage_file_with_metadata(filename, content)
    return records


def parse_usage_file_with_metadata(
    filename: str,
    content: bytes,
) -> tuple[list[UsageRecord], DatasetImportMetadata]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_usage_rows(_csv_rows(content), "CSV")
    if suffix == ".xlsx":
        return _parse_usage_rows(_xlsx_rows(content), "Excel")
    raise ValueError("Please upload a .csv or .xlsx file.")


def preview_replenishment_file(filename: str, content: bytes) -> FilePreviewResponse:
    rows = _load_rows(filename, content)
    headers = _headers_from_rows(rows)
    mappings = _preview_mappings(headers, "replenishment")
    return FilePreviewResponse(
        filename=filename,
        dataset_kind="replenishment",
        headers=headers,
        mappings=mappings,
        ai_enabled=openai_is_configured(),
    )


def preview_usage_file(filename: str, content: bytes) -> FilePreviewResponse:
    rows = _load_rows(filename, content)
    headers = _headers_from_rows(rows)
    mappings = _preview_mappings(headers, "usage")
    return FilePreviewResponse(
        filename=filename,
        dataset_kind="usage",
        headers=headers,
        mappings=mappings,
        ai_enabled=openai_is_configured(),
    )


def _csv_rows(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV file is missing a header row.")
    return [{key.strip(): value for key, value in row.items() if key} for row in reader]


def _xlsx_rows(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        raise ValueError("Excel file is missing a header row.")

    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        parsed_rows.append(
            {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
        )
    return parsed_rows


def _load_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _csv_rows(content)
    if suffix == ".xlsx":
        return _xlsx_rows(content)
    raise ValueError("Please upload a .csv or .xlsx file.")


def _headers_from_rows(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        raise ValueError("File does not contain any data rows.")
    return [header for header in rows[0].keys() if header]


def _parse_rows(
    rows: list[dict[str, str]],
    source_type: str,
) -> tuple[list[ReplenishmentRecord], DatasetImportMetadata]:
    if not rows:
        raise ValueError(f"{source_type} file does not contain any data rows.")

    headers = _headers_from_rows(rows)
    inferred_headers, sources = _resolve_headers(headers, "replenishment")
    missing = [field for field in ("sku", "name", "current_stock") if field not in inferred_headers]
    if missing:
        raise ValueError(
            f"{source_type} file is missing recognizable columns for: {', '.join(missing)}"
        )
    if "daily_demand" not in inferred_headers and not _weekly_demand_columns(headers):
        raise ValueError(
            f"{source_type} file is missing a recognizable demand column. AIR can infer demand "
            "from headers like daily demand, last week, 1wk, 2wk, 3wk, or 4wk."
        )

    data_rows = _rows_with_sku(rows, inferred_headers["sku"])
    if not data_rows:
        raise ValueError(f"{source_type} file does not contain any item rows.")

    records: list[ReplenishmentRecord] = []
    for index, row in enumerate(data_rows, start=2):
        try:
            records.append(
                ReplenishmentRecord(
                    sku=_read_text(row, inferred_headers["sku"]),
                    name=_read_text(row, inferred_headers["name"]),
                    current_stock=_read_int(row, inferred_headers["current_stock"]),
                    daily_demand=_read_daily_demand(row, headers, inferred_headers),
                    lead_time_days=_read_optional_int(
                        row, inferred_headers.get("lead_time_days"), DEFAULT_LEAD_TIME_DAYS
                    ),
                    safety_stock=_read_optional_int(
                        row, inferred_headers.get("safety_stock"), DEFAULT_SAFETY_STOCK
                    ),
                    min_order_qty=_read_optional_int(
                        row, inferred_headers.get("min_order_qty"), DEFAULT_MIN_ORDER_QTY
                    ),
                    incoming_stock=_read_optional_int(
                        row, inferred_headers.get("incoming_stock"), 0
                    ),
                    supplier_code=_read_optional_text(
                        row, inferred_headers.get("supplier_code")
                    ),
                )
            )
        except ValueError as error:
            raise ValueError(f"{source_type} row {index} contains invalid values: {error}") from error

    return records, _metadata_from_sources(sources)


def _parse_usage_rows(
    rows: list[dict[str, str]],
    source_type: str,
) -> tuple[list[UsageRecord], DatasetImportMetadata]:
    if not rows:
        raise ValueError(f"{source_type} file does not contain any data rows.")

    headers = _headers_from_rows(rows)
    inferred_headers, sources = _resolve_headers(headers, "usage")
    weekly_columns = _weekly_usage_columns(headers)
    has_transaction_usage = all(
        field in inferred_headers for field in ("sku", "usage_date", "units_used")
    )
    has_weekly_usage = "sku" in inferred_headers and bool(weekly_columns)
    has_daily_average = "sku" in inferred_headers and "units_used" in inferred_headers

    if not (has_transaction_usage or has_weekly_usage or has_daily_average):
        missing = [field for field in ("sku", "usage_date", "units_used") if field not in inferred_headers]
        raise ValueError(
            f"{source_type} file is missing recognizable usage columns for: {', '.join(missing)}. "
            "AIR also accepts weekly movement headers such as LW Mon, LW Tue, CW Mon, and CW Tue."
        )

    data_rows = _rows_with_sku(rows, inferred_headers["sku"])
    if not data_rows:
        raise ValueError(f"{source_type} file does not contain any item rows.")

    if has_weekly_usage:
        return _parse_weekly_usage_rows(data_rows, inferred_headers["sku"], weekly_columns, sources)

    if has_daily_average and "usage_date" not in inferred_headers:
        return _parse_daily_average_rows(data_rows, inferred_headers, sources)

    usage_records: list[UsageRecord] = []
    for index, row in enumerate(data_rows, start=2):
        try:
            usage_records.append(
                UsageRecord(
                    sku=_read_text(row, inferred_headers["sku"]),
                    usage_date=_parse_date(_read_text(row, inferred_headers["usage_date"])),
                    units_used=_to_float(row.get(inferred_headers["units_used"], "")),
                )
            )
        except ValueError as error:
            raise ValueError(f"{source_type} row {index} contains invalid usage values: {error}") from error

    return usage_records, _metadata_from_sources(sources)


def _parse_weekly_usage_rows(
    rows: list[dict[str, str]],
    sku_header: str,
    weekly_columns: list[tuple[str, str, int]],
    sources: dict[str, str],
) -> tuple[list[UsageRecord], DatasetImportMetadata]:
    """Turn a weekly sales snapshot into dated daily observations for trend analysis."""
    current_columns = [column for column in weekly_columns if column[1] == "current"]
    latest_current_day = _latest_active_current_day(rows, current_columns)
    today = date.today()
    current_week_start = today - timedelta(days=today.weekday())
    last_week_start = current_week_start - timedelta(days=7)

    usage_records: list[UsageRecord] = []
    for row_index, row in enumerate(rows, start=2):
        try:
            sku = _read_text(row, sku_header)
            for header, week, weekday_index in weekly_columns:
                if week == "current" and (
                    latest_current_day is None or weekday_index > latest_current_day
                ):
                    continue
                raw_value = row.get(header, "")
                if _is_blank(raw_value):
                    continue
                week_start = current_week_start if week == "current" else last_week_start
                usage_records.append(
                    UsageRecord(
                        sku=sku,
                        usage_date=week_start + timedelta(days=weekday_index),
                        units_used=max(_to_float(raw_value), 0.0),
                    )
                )
        except ValueError as error:
            raise ValueError(
                f"Weekly usage row {row_index} contains invalid values: {error}"
            ) from error

    if not usage_records:
        raise ValueError("Weekly usage columns were found, but they do not contain any usable sales data.")
    sources = {**sources, "usage_date": "heuristic", "units_used": "heuristic"}
    return usage_records, _metadata_from_sources(sources)


def _parse_daily_average_rows(
    rows: list[dict[str, str]],
    inferred_headers: dict[str, str],
    sources: dict[str, str],
) -> tuple[list[UsageRecord], DatasetImportMetadata]:
    """Use a supplied daily average when a report does not include individual dates."""
    usage_records: list[UsageRecord] = []
    for row_index, row in enumerate(rows, start=2):
        try:
            usage_records.append(
                UsageRecord(
                    sku=_read_text(row, inferred_headers["sku"]),
                    usage_date=date.today(),
                    units_used=max(_to_float(row.get(inferred_headers["units_used"], "")), 0.0),
                )
            )
        except ValueError as error:
            raise ValueError(
                f"Daily average row {row_index} contains invalid values: {error}"
            ) from error

    sources = {**sources, "usage_date": "heuristic"}
    return usage_records, _metadata_from_sources(sources)


def _rows_with_sku(rows: list[dict[str, str]], sku_header: str) -> list[dict[str, str]]:
    """Exclude export totals and separator rows that do not represent an item."""
    return [row for row in rows if not _is_blank(row.get(sku_header))]


def _infer_headers(
    headers: list[str],
    alias_map: dict[str, set[str]] = HEADER_ALIASES,
) -> dict[str, str]:
    matches: dict[str, HeaderMatch] = {}
    for header in headers:
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for field_name, aliases in alias_map.items():
            score = _header_score(normalized, aliases)
            if score <= 0:
                continue
            existing = matches.get(field_name)
            if existing is None or score > existing.score:
                matches[field_name] = HeaderMatch(field_name, header, score)
    return {field: match.source_header for field, match in matches.items() if match.score >= 0.72}


def _resolve_headers(
    headers: list[str],
    dataset_kind: str,
) -> tuple[dict[str, str], dict[str, str]]:
    alias_map = HEADER_ALIASES if dataset_kind == "replenishment" else USAGE_HEADER_ALIASES
    local_candidates = _infer_headers(headers, alias_map)
    semantic_candidates = infer_semantic_headers(headers, dataset_kind, local_candidates)
    if not semantic_candidates:
        return local_candidates, {field: "heuristic" for field in local_candidates}

    resolved = dict(local_candidates)
    resolved.update({key: value for key, value in semantic_candidates.items() if value})
    sources = {field: "heuristic" for field in local_candidates}
    for key, value in semantic_candidates.items():
        if value:
            sources[key] = "openai" if local_candidates.get(key) != value else "heuristic"
    return resolved, sources


def _preview_mappings(headers: list[str], dataset_kind: str) -> list[FileMappingPreview]:
    resolved, sources = _resolve_headers(headers, dataset_kind)
    weekly_columns = _weekly_usage_columns(headers) if dataset_kind == "usage" else []
    field_names = (
        [
            "sku",
            "name",
            "current_stock",
            "daily_demand",
            "lead_time_days",
            "safety_stock",
            "min_order_qty",
            "incoming_stock",
            "supplier_code",
        ]
        if dataset_kind == "replenishment"
        else ["sku", "usage_date", "units_used"]
    )
    previews: list[FileMappingPreview] = []
    for field in field_names:
        header = resolved.get(field)
        if dataset_kind == "usage" and weekly_columns and field in {"usage_date", "units_used"}:
            header = "Weekly movement columns (LW / CW by day)"
            sources = {**sources, field: "heuristic"}
        previews.append(
            FileMappingPreview(
                field=field,
                header=header,
                source=sources.get(field, "unmapped") if header else "unmapped",
            )
        )
    return previews


def _metadata_from_sources(sources: dict[str, str]) -> DatasetImportMetadata:
    mapped_with_ai = any(source == "openai" for source in sources.values())
    if not sources:
        return DatasetImportMetadata()
    return DatasetImportMetadata(
        mapping_origin="openai" if mapped_with_ai else "heuristic",
        mapped_with_ai=mapped_with_ai,
    )


def _header_score(normalized_header: str, aliases: set[str]) -> float:
    if normalized_header in aliases:
        return 1.0
    if any(alias in normalized_header or normalized_header in alias for alias in aliases):
        return 0.9
    return max(SequenceMatcher(None, normalized_header, alias).ratio() for alias in aliases)


def _weekly_demand_columns(headers: list[str]) -> list[tuple[str, int]]:
    weekly_headers: list[tuple[str, int]] = []
    for header in headers:
        normalized = _normalize_header(header)
        for pattern, days in WEEKLY_DEMAND_PATTERNS:
            if pattern.match(normalized):
                weekly_headers.append((header, days))
                break
    return weekly_headers


def _weekly_usage_columns(headers: list[str]) -> list[tuple[str, str, int]]:
    columns: list[tuple[str, str, int]] = []
    for header in headers:
        match = WEEKLY_USAGE_PATTERN.match(_normalize_header(header))
        if not match:
            continue
        week = "last" if match.group("week") in {"lw", "lastweek"} else "current"
        columns.append((header, week, WEEKDAY_INDEXES[match.group("day")]))
    return sorted(columns, key=lambda column: (column[1] == "current", column[2]))


def _latest_active_current_day(
    rows: list[dict[str, str]],
    current_columns: list[tuple[str, str, int]],
) -> int | None:
    active_days: list[int] = []
    for header, _, weekday_index in current_columns:
        for row in rows:
            raw_value = row.get(header, "")
            if _is_blank(raw_value):
                continue
            try:
                if _to_float(raw_value) != 0:
                    active_days.append(weekday_index)
                    break
            except ValueError:
                continue
    return max(active_days) if active_days else None


def _read_daily_demand(
    row: dict[str, str],
    headers: list[str],
    inferred_headers: dict[str, str],
) -> float:
    daily_header = inferred_headers.get("daily_demand")
    if daily_header:
        return max(_to_float(row.get(daily_header, "")), 0.0)

    weekly_columns = _weekly_demand_columns(headers)
    if not weekly_columns:
        raise ValueError("No recognizable demand column was found.")

    total_units = 0.0
    total_days = 0
    for header, days in weekly_columns:
        raw_value = row.get(header, "")
        if _is_blank(raw_value):
            continue
        total_units += _to_float(raw_value)
        total_days += days

    if total_days == 0:
        raise ValueError("Recognizable demand columns were found, but they are empty.")

    daily_demand = total_units / total_days
    return max(daily_demand, 0.0)


def _read_text(row: dict[str, str], header: str) -> str:
    value = str(row.get(header, "")).strip()
    if not value:
        raise ValueError(f"Column '{header}' is blank.")
    return value


def _read_int(row: dict[str, str], header: str) -> int:
    return int(_to_float(row.get(header, "")))


def _read_optional_int(row: dict[str, str], header: str | None, default: int) -> int:
    if not header:
        return default
    raw_value = row.get(header, "")
    if _is_blank(raw_value):
        return default
    return int(_to_float(raw_value))


def _read_optional_text(row: dict[str, str], header: str | None) -> str | None:
    if not header:
        return None
    value = str(row.get(header, "")).strip()
    return value or None


def _to_float(value: str | None) -> float:
    if value is None or _is_blank(value):
        raise ValueError("A required numeric value is blank.")
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in {"", "-", ".", "-."}:
        raise ValueError(f"Could not read numeric value '{value}'.")
    return float(cleaned)


def _parse_date(value: str) -> date:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not read date '{value}'.")


def _is_blank(value: str | None) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(header: str) -> str:
    lowered = header.strip().lower().replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", lowered)
