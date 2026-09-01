import pytest
from fastapi.testclient import TestClient
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.main import app, store
from app.models import BusinessAssumptions


client = TestClient(app)


@pytest.fixture(autouse=True)
def reset_store_state() -> None:
    store.save_records("test-reset", [])
    store.save_usage_records("usage-reset", [])
    store.save_assumptions(BusinessAssumptions())


def test_dashboard_contains_import_button() -> None:
    response = client.get("/")

    assert response.status_code == 200
    assert "Analyze Inventory" in response.text
    assert "Analyze Usage Trends" in response.text
    assert "Preview Inventory Columns" in response.text
    assert "Preview Usage Columns" in response.text
    assert "Questions Before Final Verdict" in response.text
    assert "anything else you would like it to know" in response.text.lower()
    assert "Export Verdict Workbook" in response.text
    assert "Reset Saved Answers" in response.text
    assert 'type="file"' in response.text
    assert "weekly LW/CW sales report" in response.text


def test_ingest_file_endpoint_accepts_csv_upload() -> None:
    content = "\n".join(
        [
            "sku,name,current_stock,daily_demand,lead_time_days,safety_stock,min_order_qty",
            "SKU-3001,Imported Item,6,2,4,5,10",
        ]
    )

    response = client.post(
        "/datasets/ingest-file",
        data={"source": "test-upload"},
        files={"file": ("import.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["ingested_records"] == 1


def test_preview_file_endpoint_returns_mapping_preview() -> None:
    content = "\n".join(
        [
            "Item Code,Description,On Hand,On Order,1wk,Lead Time,MOQ",
            "SKU-3001,Imported Item,6,4,14,5,10",
        ]
    )

    response = client.post(
        "/datasets/preview-file",
        files={"file": ("import.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["dataset_kind"] == "replenishment"
    assert any(item["field"] == "incoming_stock" and item["header"] == "On Order" for item in data["mappings"])


def test_preview_usage_file_endpoint_returns_mapping_preview() -> None:
    content = "\n".join(
        [
            "Item Code,Sales Date,Units Sold",
            "SKU-3001,2026-08-20,4",
        ]
    )

    response = client.post(
        "/datasets/preview-usage-file",
        files={"file": ("usage.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["dataset_kind"] == "usage"
    assert any(item["field"] == "usage_date" and item["header"] == "Sales Date" for item in data["mappings"])


def test_ingest_file_endpoint_accepts_excel_upload() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "sku",
            "name",
            "current_stock",
            "daily_demand",
            "lead_time_days",
            "safety_stock",
            "min_order_qty",
        ]
    )
    sheet.append(["SKU-4001", "Excel Item", 8, 2.5, 5, 6, 12])

    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)

    response = client.post(
        "/datasets/ingest-file",
        data={"source": "excel-upload"},
        files={
            "file": (
                "import.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["ingested_records"] == 1


def test_ingest_usage_file_endpoint_accepts_usage_upload() -> None:
    content = "\n".join(
        [
            "Item Code,Sales Date,Units Sold",
            "SKU-3001,2026-08-20,4",
            "SKU-3001,2026-08-21,6",
        ]
    )

    response = client.post(
        "/datasets/ingest-usage-file",
        data={"source": "usage-upload"},
        files={"file": ("usage.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    assert response.json()["ingested_records"] == 2


def test_recommendations_wait_for_answers_before_final_verdict() -> None:
    content = "\n".join(
        [
            "Item Code,Description,On Hand,Daily Usage",
            "SKU-5001,Fresh Basil,6,2",
        ]
    )

    upload_response = client.post(
        "/datasets/ingest-file",
        data={"source": "question-test"},
        files={"file": ("inventory.csv", content, "text/csv")},
    )

    assert upload_response.status_code == 200

    response = client.get("/recommendations")
    payload = response.json()

    assert response.status_code == 200
    assert payload["final_verdict_ready"] is False
    assert payload["verdict_origin"] == "analysis"
    assert payload["inventory_mapping_origin"] == "heuristic"
    assert payload["usage_mapping_origin"] == "none"
    assert payload["recommendations"] == []
    assert any(question["id"] == "arrival_days" for question in payload["pending_questions"])
    assert all(question["id"] != "herb_spoilage_days" for question in payload["pending_questions"])


def test_saving_assumptions_unlocks_final_verdict() -> None:
    response = client.post(
        "/analysis/assumptions",
        json={
            "shipping_days_per_week": 5,
            "arrival_days": ["Monday", "Wednesday", "Friday"],
            "default_spoilage_days": 7,
            "produce_spoilage_days": 4,
            "herb_spoilage_days": 2,
            "additional_notes": "Some herbs often arrive spoiled and get claimed.",
        },
    )

    assert response.status_code == 200
    assert response.json()["final_verdict_ready"] is True
    assert response.json()["assumptions"]["additional_notes"].startswith("Some herbs")


def test_saving_partial_assumptions_keeps_previously_saved_answers() -> None:
    client.post(
        "/analysis/assumptions",
        json={
            "shipping_days_per_week": 5,
            "arrival_days": ["Monday", "Wednesday", "Friday"],
        },
    )

    response = client.post(
        "/analysis/assumptions",
        json={
            "default_spoilage_days": 4,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["assumptions"]["shipping_days_per_week"] == 5
    assert payload["assumptions"]["arrival_days"] == ["Monday", "Wednesday", "Friday"]
    assert payload["assumptions"]["default_spoilage_days"] == 4


def test_reset_assumptions_clears_saved_memory() -> None:
    client.post(
        "/analysis/assumptions",
        json={
            "shipping_days_per_week": 5,
            "arrival_days": ["Monday", "Wednesday", "Friday"],
            "default_spoilage_days": 4,
            "additional_notes": "Old planning memory.",
        },
    )

    response = client.post("/analysis/assumptions/reset")

    assert response.status_code == 200
    payload = response.json()
    assert payload["assumptions"]["shipping_days_per_week"] is None
    assert payload["assumptions"]["arrival_days"] == []
    assert payload["assumptions"]["default_spoilage_days"] is None
    assert payload["assumptions"]["additional_notes"] == ""


def test_export_requires_required_answers_first() -> None:
    content = "\n".join(
        [
            "Item Code,Description,On Hand,Daily Usage",
            "SKU-5001,Fresh Basil,6,2",
        ]
    )

    client.post(
        "/datasets/ingest-file",
        data={"source": "question-test"},
        files={"file": ("inventory.csv", content, "text/csv")},
    )

    response = client.get("/recommendations/export")

    assert response.status_code == 400
    assert "still needs required planning answers" in response.json()["detail"]


def test_export_returns_excel_file_when_verdict_is_ready() -> None:
    content = "\n".join(
        [
            "Item Code,Description,On Hand,Daily Usage",
            "SKU-6001,Canned Black Beans,3,2",
        ]
    )

    client.post(
        "/datasets/ingest-file",
        data={"source": "export-test"},
        files={"file": ("inventory.csv", content, "text/csv")},
    )
    client.post(
        "/analysis/assumptions",
        json={
            "shipping_days_per_week": 5,
            "arrival_days": ["Monday", "Wednesday", "Friday"],
            "additional_notes": "Export test context.",
        },
    )

    response = client.get("/recommendations/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Verdict Detail", "Assumptions & Inputs"]
    assert workbook["Summary"]["A1"].value == "AIR Verdict Summary"
    assert workbook["Summary"]["D4"].value == "SKU Number"
    assert workbook["Verdict Detail"]["A1"].value == "SKU Number"
    assert workbook["Verdict Detail"]["A2"].value == "SKU-6001"
    assert workbook["Assumptions & Inputs"]["A1"].value == "AIR Planning Assumptions"
    assert workbook["Assumptions & Inputs"]["B4"].value == 5
    assert workbook["Assumptions & Inputs"]["B5"].value == "Monday, Wednesday, Friday"
    assert workbook["Assumptions & Inputs"]["B9"].value == "Export test context."


def test_recommendations_show_usage_mapping_origin_when_usage_file_is_loaded() -> None:
    inventory_content = "\n".join(
        [
            "Item Code,Description,On Hand,Daily Usage",
            "SKU-7001,Usage Driven Item,30,2",
        ]
    )
    usage_content = "\n".join(
        [
            "Item Code,Sales Date,Units Sold",
            "SKU-7001,2026-08-20,8",
            "SKU-7001,2026-08-21,9",
        ]
    )

    client.post(
        "/datasets/ingest-file",
        data={"source": "origin-test"},
        files={"file": ("inventory.csv", inventory_content, "text/csv")},
    )
    client.post(
        "/datasets/ingest-usage-file",
        data={"source": "origin-usage"},
        files={"file": ("usage.csv", usage_content, "text/csv")},
    )
    client.post(
        "/analysis/assumptions",
        json={
            "shipping_days_per_week": 5,
            "arrival_days": ["Monday", "Wednesday", "Friday"],
        },
    )

    response = client.get("/recommendations")

    assert response.status_code == 200
    payload = response.json()
    assert payload["verdict_origin"] == "analysis"
    assert payload["inventory_mapping_origin"] == "heuristic"
    assert payload["usage_mapping_origin"] == "heuristic"
